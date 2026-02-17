#!/usr/bin/env python3
"""
Extract SLAS seniority list XLSX files into JSON matching the 2026 schema.

Usage:
    python extract_xlsx.py                    # Extract all xlsx files
    python extract_xlsx.py 2021_grade_i       # Extract single file
    python extract_xlsx.py --year 2024        # Extract all grades for a year
"""

import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

RAW_DIR = Path(__file__).parent.parent / "raw_data"
DATA_DIR = Path(__file__).parent.parent / "data"

GRADE_CONFIG = {
    "sp_grade": {
        "title": "List of Special Grade Officers",
        "grade": "Special Grade",
        "has_serial_and_seniority": True,
        "col_count": 11,
        "date_fields": [
            "dateOfBirth",
            "dateOfEntryToGradeIII",
            "dateOfPromotionToGradeII",
            "dateOfPromotionToGradeI",
            "dateOfPromotionToGradeSP",
        ],
    },
    "grade_i": {
        "title": "List of Grade I SLAS Officers",
        "grade": "Grade I",
        "has_serial_and_seniority": True,
        "col_count": 10,
        "date_fields": [
            "dateOfBirth",
            "dateOfEntryToGradeIII",
            "dateOfPromotionToGradeII",
            "dateOfPromotionToGradeI",
        ],
    },
    "grade_ii": {
        "title": "List of Grade II SLAS Officers",
        "grade": "Grade II",
        "has_serial_and_seniority": False,
        "col_count": 8,
        "date_fields": [
            "dateOfBirth",
            "dateOfEntryToGradeIII",
            "dateOfPromotionToGradeII",
        ],
    },
    "grade_iii": {
        "title": "List of Grade III SLAS Officers",
        "grade": "Grade III",
        "has_serial_and_seniority": False,
        "col_count": 7,
        "date_fields": [
            "dateOfBirth",
            "dateOfEntryToGradeIII",
        ],
    },
}

CONTACT_INFO = (
    "If there are any corrections to be done please contact "
    "SLAS Branch. (TP 0112698605)"
)

FILE_NUM_RE = re.compile(r"^75/10/\d{3,5}$")


def fmt_date(val):
    """Convert a cell value to M/D/YYYY string."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return f"{val.month}/{val.day}/{val.year}"
    s = str(val).strip()
    if not s:
        return ""
    return s


def clean(val):
    """Clean a cell value into a string, collapsing whitespace."""
    if val is None:
        return ""
    s = str(val).replace("\n", " ").replace("\r", " ")
    return re.sub(r"\s+", " ", s).strip()


def detect_columns(ws):
    """Find the header row and map column names to indices.

    Returns (header_row, col_map) where col_map maps semantic names
    to 0-based column indices.
    """
    header_row = None
    for r in range(1, 15):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and "File" in str(v):
                header_row = r
                break
        if header_row:
            break

    if header_row is None:
        return None, None

    # Read all header cells
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        headers.append((c, str(v).strip().lower().replace("\n", " ") if v else ""))

    col_map = {}
    for c, h in headers:
        if not h:
            continue
        hl = h.lower()
        if "serial" in hl:
            col_map["serial"] = c
        elif "seniority" in hl or "sen." in hl and "serial" not in col_map:
            col_map.setdefault("seniority", c)
        elif "file" in hl:
            col_map["file"] = c
        elif hl == "name":
            col_map["name"] = c
        elif "post" in hl and "name" not in hl:
            col_map["post"] = c
        elif "work" in hl or "place" in hl:
            col_map["workplace"] = c
        elif "birth" in hl:
            col_map["dob"] = c
        elif "entry" in hl:
            col_map["entry_giii"] = c
        elif "promo" in hl and "ii" in hl and "promo_gii" not in col_map and "i" == hl.split()[-1]:
            # Promotion to Grade I (last word is just "i")
            col_map["promo_gi"] = c
        elif "promo" in hl and ("sp" in hl or "special" in hl):
            col_map["promo_sp"] = c
        elif "promo" in hl and "ii" in hl and "promo_gii" not in col_map:
            col_map["promo_gii"] = c
        elif "promo" in hl and "i" in hl:
            # Catch remaining promotion columns by order
            if "promo_gii" not in col_map:
                col_map["promo_gii"] = c
            elif "promo_gi" not in col_map:
                col_map["promo_gi"] = c
            elif "promo_sp" not in col_map:
                col_map["promo_sp"] = c

    # For Grade II/III sheets that use "Sen. No." for both serial and seniority
    if "serial" not in col_map and "seniority" in col_map:
        col_map["serial"] = col_map["seniority"]

    return header_row, col_map


def extract_xlsx(year, grade_key):
    """Extract a single xlsx into the JSON schema."""
    config = GRADE_CONFIG[grade_key]
    xlsx_path = RAW_DIR / f"{year}_{grade_key}.xlsx"

    if not xlsx_path.exists():
        print(f"  NOT FOUND: {xlsx_path}")
        return None

    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active

    header_row, col_map = detect_columns(ws)
    if header_row is None or not col_map:
        print(f"  ERROR: cannot detect columns in {xlsx_path.name}")
        wb.close()
        return None

    data_start = header_row + 1

    # Map date_fields to col_map keys
    date_field_to_col = {
        "dateOfBirth": "dob",
        "dateOfEntryToGradeIII": "entry_giii",
        "dateOfPromotionToGradeII": "promo_gii",
        "dateOfPromotionToGradeI": "promo_gi",
        "dateOfPromotionToGradeSP": "promo_sp",
    }

    officers = []
    for r in range(data_start, ws.max_row + 1):
        # Read serial number
        serial_val = ws.cell(row=r, column=col_map.get("serial", 1)).value
        if serial_val is None or not isinstance(serial_val, (int, float)):
            continue

        # Read file number
        fn_val = ws.cell(row=r, column=col_map["file"]).value
        fn = str(fn_val).strip() if fn_val else ""
        if not fn or not FILE_NUM_RE.match(fn):
            continue

        sn = int(serial_val)
        sen = int(ws.cell(row=r, column=col_map.get("seniority", col_map["serial"])).value or sn)

        # Read workplace — some xlsx files have an extra empty column
        # between post and workplace, so data may be in either column
        wp = ""
        if "workplace" in col_map:
            wp = clean(ws.cell(row=r, column=col_map["workplace"]).value)
            if not wp:
                # Check the column just before (empty-header column)
                alt_col = col_map["workplace"] - 1
                if alt_col > col_map["post"]:
                    wp = clean(ws.cell(row=r, column=alt_col).value)

        officer = {
            "serialNo": sn,
            "currentSeniorityNo": sen,
            "fileNumber": fn,
            "name": clean(ws.cell(row=r, column=col_map["name"]).value),
            "presentPost": clean(ws.cell(row=r, column=col_map["post"]).value),
            "presentWorkPlace": wp,
        }

        for field in config["date_fields"]:
            col_key = date_field_to_col[field]
            if col_key in col_map:
                officer[field] = fmt_date(ws.cell(row=r, column=col_map[col_key]).value)
            else:
                officer[field] = ""

        officers.append(officer)

    wb.close()

    result = {
        "title": config["title"],
        "asAt": f"01.01.{year}",
        "grade": config["grade"],
        "contactInfo": CONTACT_INFO,
        "officers": officers,
    }
    return result


def validate(result, year, grade_key):
    """Validate and report on extraction quality."""
    if not result:
        return False

    officers = result["officers"]
    if not officers:
        print(f"  WARNING: No officers extracted")
        return False

    no_name = sum(1 for o in officers if not o["name"])
    no_post = sum(1 for o in officers if not o["presentPost"])
    no_wp = sum(1 for o in officers if not o["presentWorkPlace"])

    issues = []
    if no_name:
        issues.append(f"{no_name} missing names")
    if no_post:
        issues.append(f"{no_post} missing posts")
    if no_wp:
        issues.append(f"{no_wp} missing workplaces")

    for field in GRADE_CONFIG[grade_key]["date_fields"]:
        missing = sum(1 for o in officers if not o.get(field))
        if missing > len(officers) * 0.1:
            issues.append(f"{missing}/{len(officers)} missing {field}")

    status = "OK" if not issues else "WARN"
    print(f"  {status}: {len(officers)} officers", end="")
    if issues:
        print(f"  ({', '.join(issues)})", end="")
    print()
    return True


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    years = ["2024", "2023", "2022", "2021"]
    grades = ["sp_grade", "grade_i", "grade_ii", "grade_iii"]

    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if arg.startswith("--year"):
            year = sys.argv[2] if len(sys.argv) > 2 else arg.split("=")[1]
            years = [year]
        elif "_" in arg:
            parts = arg.split("_", 1)
            years = [parts[0]]
            grades = [parts[1]]

    for year in years:
        print(f"\n{'='*50}")
        print(f"  Year {year}")
        print(f"{'='*50}")

        for grade_key in grades:
            print(f"\n  --- {year}_{grade_key} ---")
            result = extract_xlsx(year, grade_key)

            if result and validate(result, year, grade_key):
                output_path = DATA_DIR / f"{year}_{grade_key}.json"
                with open(output_path, "w", encoding="utf-8") as f:
                    json.dump(result, f, indent=2, ensure_ascii=False)
                print(f"  Saved: {output_path.name}")
            else:
                print(f"  FAILED: {year}_{grade_key}")


if __name__ == "__main__":
    main()
